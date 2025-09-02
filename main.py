from tkinter import *
from tkinter import ttk
from tkinter import filedialog
from tkinter import messagebox
import openpyxl
import secrets
import xlrd
import os


def start_imei_check():
    path = filedialog.askopenfilename(title="Select XLS file",
                                        filetypes=[("Excel files", "*.xls")])
    if path:
        workbook = xlrd.open_workbook(path)
        sheet = workbook.sheet_by_index(0)

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = sheet.name

    for row in range(sheet.nrows):
        for col in range(sheet.ncols):
            ws.cell(row=row+1, column=col+1, value=sheet.cell_value(row, col))

    ws.insert_cols(6)
    ws.cell(row=1, column=6, value="imei")
    list_of_imeis = get_imeis_from_file(ws)
    imei_managed_list = [correct_imei_check_digit(imei) for imei in list_of_imeis]

    new_col_index = 6

    for row_num, val in enumerate(imei_managed_list, start=2):
        ws.cell(row=row_num, column=new_col_index, value=val)
    ws.column_dimensions['E'].hidden = True
    new_file_path = os.path.splitext(path)[0] + f"_{secrets.token_hex(nbytes=8)}.xlsx"
    wb.save(new_file_path)
    messagebox.showinfo("Success", f"File saved as {new_file_path}")


def get_imeis_from_file(ws) -> list | None:
    col_index = 5
    values = [ws.cell(row=i, column=col_index).value for i in range(2, ws.max_row+1)]
    if not values:
        return None
    return values


def correct_imei_check_digit(imei: str) -> str | None:
    if not isinstance(imei, str) or not imei.isnumeric() or len(imei) != 15:
        return None
    
    total_sum = 0
    alternate = True
    
    for i in range(len(imei) - 2, -1, -1):
        digit = int(imei[i])
        
        if alternate:
            doubled_digit = digit * 2
            if doubled_digit > 9:
                total_sum += doubled_digit - 9
            else:
                total_sum += doubled_digit
        else:
            total_sum += digit
        alternate = not alternate

    control_digit = (10 - (total_sum % 10)) % 10
    original_check_digit = int(imei[14])

    if control_digit == original_check_digit:
        return imei
    else:
        return imei[:-1] + str(control_digit)


def close_window():
    root.destroy()

root = Tk()
root.title("IMEI")
root.geometry(("300x150"))
readme_label = ttk.Label(text="Select XLS file downloaded from Partner")
readme_label.pack(pady=20)
btn = ttk.Button(root, text="Choose File", command=start_imei_check)
btn.pack()
btn_exit = ttk.Button(root, text="Exit", command=close_window)
btn_exit.pack()
root.mainloop()

